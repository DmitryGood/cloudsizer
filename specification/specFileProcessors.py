__author__ = 'slash'
from exceptions import NotImplementedError
import xlrd
import openpyxl
from exceptions import ValueError


class AbstractXLSSheetProcessor(object):
    ''' abstract class to process XLS files
    '''
    def __init__(self, filename):
        self.open_workbook(filename)

    def open_workbook(self, filename):
        raise ('Subclasses must override open_workbook()')

    def sheet_names(self):
        raise ('Subclasses must override sheet_names()')

    def sheet_by_index(self, index):
        raise ('Subclasses must override sheet_by_index()')

    def cell(self, row, col):
        raise ('Subclasses must override cell()')

    def isCellEmpty(self, row, col):
        raise ('Subclasses must override isCellEmpty()')

    def max_row(self):
        raise ('Subclasses must override max_row()')

    def max_col(self):
        raise ('Subclasses must override max_col()')

    def row(self, row):
        raise ('Subclasses must override row()')

class xlsSheetProcessor(AbstractXLSSheetProcessor):
    def open_workbook(self, filename):
        if (filename != None):
            self.book = xlrd.open_workbook(filename)
        else:
            raise ValueError('File name cannot be empty')
        return self

    def sheet_names(self):
        return self.book.sheet_names()

    def sheet_by_index(self, index):
        self.sheet = self.book.sheet_by_index(0)
        return self

    def cell(self, row, col):
        return self.sheet.cell(row, col)

    def isCellEmpty(self, row, col):
        type = self.sheet.cell(row, col).ctype
        if (type == 0 or type == 6):      # 0 - XL_CELL_EMPTY, 6 - XL_CELL_BLANK)
            return True
        return False

    def max_row(self):
        return self.sheet.nrows

    def max_col(self):
        return self.sheet.ncols

    def row(self, row):
        return self.sheet.row(row)

class xlsxSheetProcessor(AbstractXLSSheetProcessor):
    def open_workbook(self, filename):
        if (filename != None):
            self.book = openpyxl.load_workbook(filename)
        else:
            raise ValueError('File name cannot be empty')
        return self

    def sheet_names(self):
        return self.book.get_sheet_names()

    def sheet_by_index(self, index):
        self.sheet = self.book.worksheets[index]
        return self

    def cell(self, row, col):
        return self.sheet.cell(row=row+1, column=col+1)

    def isCellEmpty(self, row, col):
        if (not self.sheet.cell(row=row+1, column=col+1).value ):
            return True
        return False

    def max_row(self):
        return self.sheet.max_row+1

    def max_col(self):
        return self.sheet.max_column+1

    def row(self, row):
        return self.sheet.row(row+1)


