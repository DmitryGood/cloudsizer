__author__ = 'slash'
from openpyxl import load_workbook
from exceptions import ValueError
from sqlalchemy.orm.exc import NoResultFound
from model_cloudcalc import Base, Gpl, Gpl_line, Category


class GplFactory():
    ''' Class to process all actions with loading GPL into database
    '''
    # --------------- Static data --------------
    # categories for columns
    CAT_PN = 1
    CAT_NAME = 2
    CAT_PRICE = 3
    CAT_BUNDLE = 4

    # Look for header line
    column_headers = ['Product', 'Product Description', 'Price in USD', 'Item Identifier']
    column_categories = [CAT_PN, CAT_NAME, CAT_PRICE, CAT_BUNDLE]

    # product categories
    PROD_CPU = 'CPU'
    PROD_MEM = 'MEM'
    PROD_HDD = 'HDD'
    PROD_SSD = 'SDD'
    PROD_BUNDL = 'BUND'


    # filter for product categories
    product_filters = {
        PROD_CPU : ['A01-','N20-X0', 'UCS-CPU-E3', 'UCS-CPU-E5', 'UCS-CPU-E7', 'UCS-CPUE5', 'UCSCPU'],
        PROD_MEM : ['A02-','N01-M3', 'UCS-ML', 'UCS-MR', 'UCS-MU', 'UCS-SPL-M', 'UCS-SPM-M16', 'UCS-SPM-M32', ],
        PROD_HDD : ['A03-','R200-D', 'UCS-EZ-300', 'UCS-EZ7-300GB', 'UCS-HD', 'UCS-SP-1P2T', 'UCS-SPL-1P2T', 'UCS-SPL-D1P2T', 'UCSHDD'],
        PROD_SSD : ['N20-D0','UCS-C3X60', 'UCS-SD', 'UCSSD', 'UCSSSD'],
        PROD_BUNDL : ['C880-2T', 'C880-6T', 'UCS-CX-B', 'UCS-EZ7-B', 'UCS-SA', 'UCS-SL-CPA',
                  'UCS-SM', 'UCS-SP-B', 'UCS-SP-C',
                  'UCS-SP7-SRB200', 'UCS-SP7SRB200', 'UCS-SPL-B', 'UCS-SPL-C', 'UCS-SPM-B', 'UCS-SPM-C',
                  'UCS-SPR-C', 'UCS-SR-B', 'UCS-VDI-C', 'UCSB-EZ-UC-', 'UCSCDBUNC2', 'UCSEZ', 'UCSME-',
                  'UCSSP', 'UCUCS-EZ-B', 'UCUCS-EZ-C', 'UCUCSEZ-C']
    }

    # ----------------- Static data end ------------

    @staticmethod
    def define_category(pn):
        '''
        :param pn: Par number
        :return: Category for partnumber of None
        '''
        category = None
        for key in GplFactory.product_filters:
            for filter in GplFactory.product_filters[key]:
                if (pn.startswith(filter)):
                    return key
        return None

    @staticmethod
    # convert price value to integer
    def convertPriceToInt(priceStr):
        if (priceStr and priceStr.startswith('$ ') and priceStr[-3] == '.'):
            temp = priceStr[1: -3].strip()
            result = temp.replace(',', '')
            return int(result)
        return 0

    def find_header(self, sheet):    # Throws an exception if header not found
        '''
        :param sheet: sheet object for specification
        :return: dictionary {'firstline': <first line of pricelist>, 'matching' : <categories to column mathching map>
        '''
        line = 1
        nrows = sheet.max_row
        ncols = sheet.max_column
        found = False
        result={}
        while (not found and line < nrows):
            for col in range(1, ncols+1):
                cell = sheet.cell(column=col, row=line)
                cellv = cell.value
                if cellv in self.column_headers:
                    index = self.column_headers.index(cellv)
                    result[self.column_categories[index]] = col
                    found = True
            # Increase line num
            line +=1
        # After the loop line = first line of the specification
        # If we here - the cycle has been stopped
        if (line >=nrows or len(result) != len(self.column_headers)):
            # We haven't found header or all 4 columns
            raise ValueError("Didn't find header or columns in file")

        return {'firstline': line, 'matching' : result}

    # Constructor - create and initialize object from file
    def __init__(self,filename):
        try:
            self.wb = load_workbook(filename)
            self.sheet = self.wb.worksheets[0]
        except:
            raise ValueError("Can't open file")
        try:
            temp = self.find_header(self.sheet)
        except ValueError as e:
            raise e
        self.firstline = temp['firstline']
        self.matching = temp['matching']
        return      # We completely ready to upload specification

    # Create database categories if isn't exist or retrieve them
    def get_product_categories(self, session):
        '''
        :param session: database session
                look for categories in database and create if doesn't exist
        :return: dictionary with {<product_cat> : <category class from db>)
        '''
        result = {self.PROD_CPU: None,
                  self.PROD_MEM: None,
                  self.PROD_HDD: None,
                  self.PROD_SSD: None,
                  self.PROD_BUNDL: None}
        for key in result:
            try:
                # find category object in database
                cat = session.query(Category).filter(Category.name == result[key]).one()
            except NoResultFound:
                # If category isn't exist - create it
                cat = Category(name = key, filter_list= self.product_filters[key])
                session.add(cat)
                session.commit()
            result[key]=cat             # put category object to result dictionary
        return result

    # process GPL file line-by-line and load lines to database
    def load_gpl_to_db(self, session, gpl):
        '''
        :param session: database session
        :param gpl: GPL object in database - represents GPL file
        :return:    rows in gpl_lines table
                    user categories from category table
        '''
        categories = self.get_product_categories(session)
        nrows = self.sheet.max_row
        total = 0
        found = 0
        for row in range(self.firstline, nrows+1):
            pn = self.sheet.cell(row=row, column=self.matching[self.CAT_PN]).value
            descr = self.sheet.cell(row=row, column=self.matching[self.CAT_NAME]).value
            price_val = self.sheet.cell(row=row, column=self.matching[self.CAT_PRICE]).value
            price = GplFactory.convertPriceToInt(price_val)
            t = self.sheet.cell(row=row, column=self.matching[self.CAT_BUNDLE]).value
            bundle = not (type == 'Product')
            # If Parnumber field empty - skip line
            if (pn == None or pn ==''):
                continue
            total += 1
            cat = GplFactory.define_category(pn)
            # If can't define category for string - skip line
            if (cat == None):
                continue
            line = Gpl_line(pn=pn, description = descr, price=int(price), category=categories[cat],  gpl = gpl)
            session.add(line)
            session.commit()
            found +=1
        print "--- %s lines processed, %s actual records found"%(total, found)
        return {'total' : total, 'found': found}
