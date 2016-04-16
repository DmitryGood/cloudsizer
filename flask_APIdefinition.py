
import datetime
import os

from flask import Flask, redirect, make_response
from flask_sqlalchemy import SQLAlchemy
from werkzeug import secure_filename
from flask import send_from_directory, send_file
from flask import request, jsonify, json
from sqlalchemy.orm.exc import NoResultFound

from config import WorkConfig
from specification.SpecFactory import SpecFactory
from specification.BundleFactory import BundleFactory
from model_cloudcalc import User, User_action  # database types
from users.userRegistrator import UserSession

#from project.textTokenizer import Tokenizer


# config
app = Flask(__name__, static_folder=WorkConfig.STATIC_FOLDER)
#if ('TESTING' in globals()):
#    app.config.from_object(TestConfig)
#else:
app.config.from_object(WorkConfig)
db = SQLAlchemy(app)

# Additional import
#from flask.ext.httpauth import HTTPBasicAuth
#auth = HTTPBasicAuth()

# static routes
#print "App static folder: ", app.static_folder


@app.route('/')
def index():
    print "App static folder: ", app.static_folder
    return app.send_static_file('index.html'), 200

@app.route('/<path:path>', methods=['GET'])
def static_site(path):
    return app.send_static_file(path), 200

''' ------------ API stuff begins
'''
## -------- tools -----
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in app.config['ALLOWED_EXTENSIONS']

## --------- end -------




@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    user_session = UserSession(db.session, request)         # create user session
    if file and allowed_file(file.filename):
        # get user's credentials and save him

        #ip_addr = request.remote_addr
        #user = User(None, User.USER_ROLE_USER,{'ip': ip_addr})
        #db.session.add(user)
        #db.session.commit()
        # user added - this part should be changed to registration

        # add to filename user's cookies
        filename = secure_filename(user_session.getCookie() + '_'+file.filename)             # new secure filename with cookie string
        dest_filename = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        # save specification to disk
        file.save(dest_filename)
        # calculate the hash

        print "Internal path on server: ", filename, app.config['UPLOAD_FOLDER']
        # Log upload
        log = open(os.path.join(app.config['UPLOAD_FOLDER'], '_uploads.log'), 'a+')
        log.write("%s : upload file: %s, from user: %s, IP: %s\n"%(
            datetime.datetime.now(),
            filename,
            user_session.getUserID(),
            request.remote_addr
        ))

        # Load specification to database
        try:
            spec_factory = SpecFactory(dest_filename)
            print "specification factory created"
            hash = spec_factory.uploadSpecToDatabase(db.session, user_session.getUserID())
            print "Spec hash: '%s', redirecting"%hash
            resp = make_response(redirect('#/specification/' + hash))
        except Exception as e:
            print "Exceptions happens: ", e
            resp = make_response(redirect('/'), 404)           # Add exception handler - redirect to bug report page. /#/bugreport/ par:{filename :}
    else:
        resp = make_response(redirect('/'), 404)
    user_session.setCookies(resp)
    return resp


@app.route('/api/upload', methods=['POST'])
def api_upload_file():
    file = request.files['file']
    user_session = UserSession(db.session, request)         # create user session
    try:                                # Try to get name from request
        param = request.form['data']
        jsonObject = json.loads(param)
        name = jsonObject['name']
    except:                             # set name to none if can't
        name = None
    if file and allowed_file(file.filename):
        print ('--------->>>>>> Upload file through API call:')
        # get user's credentials and save him
        #ip_addr = request.remote_addr
        #user = User(None, User.USER_ROLE_USER,{'ip': ip_addr})
        #db.session.add(user)
        #db.session.commit()
        # add to filename user's cookies
        filename = secure_filename(str(user_session.getUserID()) + '_'+file.filename)             # new secure filename with cookie string
        dest_filename = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        # save specification to disk
        file.save(dest_filename)
        # calculate the hash

        print "Internal path on server: ", filename, app.config['UPLOAD_FOLDER']
        # Log upload
        log = open(os.path.join(app.config['UPLOAD_FOLDER'], '_uploads.log'), 'a+')
        log.write("%s : upload file: %s, from user: %s, IP: %s\n"%(
            datetime.datetime.now(),
            filename,
            user_session.getUserID(),
            request.remote_addr
        ))

        # Load specification to database
        try:
            print "Spec name: ", name
            spec_factory = SpecFactory(dest_filename, specName=name)
            print "specification factory created"
            hash = spec_factory.uploadSpecToDatabase(db.session, user_session.getUserID())
            print "Spec hash: '%s', return success"%hash
            user_session.registerEvent(db.session,User_action.UPLOAD_SPEC, request, data={'filename' : dest_filename, 'hash': hash})    # register spec upload
            resp = make_response(jsonify({'result' : True, 'hash' : hash}), 200)
        except Exception as e:
            print "Exceptions happens: ", e
            resp = make_response({'result' : False, 'message' : 'Upload failed'}, 404)
    else:
        resp = make_response(jsonify({'result' : False, 'message' : 'File is not allowed'}), 404)
    user_session.setCookies(resp)
    return resp


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],
                               filename)

#@app.route('/api/specification', methods=['PUT'])
#def send_spec_to_server():
#    json_data = request.json

@app.route('/api/specification/<hash>', methods=['GET'])
def get_specification(hash):
    print "Got parameter: ", hash
    user_session = UserSession(db.session, request)         # create user session
    user_session.registerEvent(db.session, User_action.FIND_SPEC,request,data={'hash' : hash})
    result = SpecFactory.calculateSpecResources(db.session, hash)
    return jsonify({'result': True, 'data': result}), 200

@app.route('/api/registerevent', methods=['POST'])
def register_event():
    ''' Function parameter (JSON): event - type of the event
    :return: user's browser has cookies and session ID
    '''
    user_session = UserSession(db.session, request)         # create user session
    try:                                # Try to get name from request
        #param = request.json
        jsonObject = request.json
        event = jsonObject['event']
    except:                             # set name to none if can't
        event = None
        print "Event is None, user session: %s"%(user_session.getUserID())
    else:
        if (event and event != User_action.REGISTER and event != User_action.CONNECT ):
            try:
                eventData=jsonObject['eventData']
            except KeyError as e:
                eventData = None
            user_session.registerEvent(db.session,event,request,data=eventData)
            print "registering event %s, with data %s"%(event, eventData)
    resp = make_response(jsonify({'result' : True}), 200)
    user_session.setCookies(resp)
    return resp

@app.route('/api/register', methods=['POST'])
def register():
    json_data = request.json
    user = User(
        name=json_data['name']

    )
    print ("new user: ", user)
    result={}
    try:
        db.session.add(user)
        db.session.commit()
        result['result'] = True
    except:
        result['result'] = False
        result['message'] = 'this user is already registered'
    db.session.close()
    return jsonify(result)



@app.route('/api/login', methods=['POST'])
def login():
    json_data = request.json
    try:
        user = db.session.query(User).filter(User.email == json_data['email']).one()
    except NoResultFound:
        print "Login API: NoResultFound exception while DB access"
        return jsonify({'result': False})
    except:
        print "Login API: unknown exception while DB access"
        return jsonify({'result': False})

    if user and user.checkPassword(json_data['password']):
        status = True
        resp = app.make_response(jsonify({'result': status, "userLogin" : user.email, "userToken" : user.get_token_string()}))
        resp.set_cookie('userLogin',value=user.email)
        resp.set_cookie('userToken', value=user.get_token_string())
        resp.set_cookie('userRole', value=str(user.role))
        print ("Server: REST API login: ", user.email, status)
    else:
        status = False
        resp = app.make_response(jsonify({'result': status, }))
    return resp


#@app.route('/logout', methods=['GET', 'POST'])
@app.route('/api/logout', methods=['POST'])
def logout():
    resp = app.make_response(jsonify({'result': 'success'}))
    resp.set_cookie('userLogin', '', expires=0)
    resp.set_cookie('userToken', '', expires =0)
    resp.set_cookie('userRole', '', expires =0)
    print ("Logging Response: ", resp)
    return resp



''' The function returns configurations for bundle (C220 with Vmware Ent Plus is supported for now
        Returns: array of possible configs for HX-220
'''
@app.route("/api/hyperflex/<model>", methods=['POST','GET'])
def hyperflex_config(model):
    print model
    user_session = UserSession(db.session, request)  # create user session
    #basedir = os.path.abspath(os.path.dirname(__file__))
    #bf1 = BundleFactory(basedir + '/data/Hyperflex_1_upload.xlsx')
    #bf2 = BundleFactory(basedir + '/data/Hyperflex_2_upload.xlsx')
    #bf3 = BundleFactory(basedir + '/data/Hyperflex_3_upload.xlsx')

    basedir = os.path.abspath(os.path.dirname(__file__))
    print "Load Hyperflex config from directory: ", basedir

    bf1 = BundleFactory(basedir + '/data/Hyperflex_1_upload.xlsx',"1")
    bf2 = BundleFactory(basedir + '/data/Hyperflex_2_upload.xlsx',"2")
    bf3 = BundleFactory(basedir + '/data/Hyperflex_3_upload.xlsx',"3")
    bf1.extractBundle()
    bf2.extractBundle()
    bf3.extractBundle()
    result = BundleFactory.combine_bundles_config([bf1, bf2, bf3])
    resp = make_response(jsonify({'result' : True, 'data': result}), 200)
    user_session.setCookies(resp)
    return resp

''' The function returns *.xlsx file for particular configuration
    Params: {"BundleID" : ID, "bundleParams" : { "servers" : X, "memory" : Y}}
    Result: specification file
'''
from openpyxl import Workbook       # remove after refactoring
@app.route("/api/hyperflex/specification/<model>", methods=['POST','GET'])
def hyperflex_specification(model):
    #print model
    print "Start spec processing", model
    #print request
    user_session = UserSession(db.session, request)  # create user session

    jsonObject = request.json
    #print jsonObject
    bundleId = jsonObject['BundleID']
    bundleParams = jsonObject['bundleParams']
    basedir = os.path.abspath(os.path.dirname(__file__))
    bf = BundleFactory.loadBundleByID(bundleId, basedir + "/data/")
    bf.extractBundle()
    wb1 = Workbook()
    ws1 = wb1.create_sheet(title="Spec", index=0)
    mapping1 = {BundleFactory.CAT_PN : 2, BundleFactory.CAT_QUANTITY : 3, BundleFactory.CAT_PRICE : 5, BundleFactory.CAT_TOTAL : 6}
    headers1 = {BundleFactory.CAT_PN : "Part Number",
           BundleFactory.CAT_QUANTITY : "Qty",
           BundleFactory.CAT_PRICE : "Unit List Price",
           BundleFactory.CAT_TOTAL : "Total list price"}
    bf.create_sheet_header(ws1, 1, mapping1, headers1)
    row = bf.spec_to_worksheet(ws1, 2, 'BASE', mapping1, 1 , True)
    row = bf.spec_to_worksheet(ws1, row, 'OPTION', mapping1, bundleParams['servers'], True)
    row = bf.spec_to_worksheet(ws1, row, 'ADDON', mapping1, bundleParams['memory'], True)

    #filename = basedir + "/hyperflex_config/hx-" + str(user_session.getUserID()) + "-" + str(datetime.datetime.now().microsecond)+".xslx"
    dir = basedir + "/hyperflex_config/"
    file = "hx-" + str(user_session.getUserID()) + "-" + str(datetime.datetime.now().microsecond)+".xlsx"
    filename = dir + file
    print "Filename: ", filename
    print dir, file
    wb1.save(filename=filename)

    #return send_file(filename, attachment_filename="hx-" + model+".xlsx")
    return send_from_directory(dir,
                               file)


@app.route("/api/hyperflex/calculator", methods=['POST','GET'])
def hyperflex_calculator():
    user_session = UserSession(db.session, request)  # create user session

    jsonObject = request.json
    # print jsonObject
    bundleId = jsonObject['BundleID']
    bundleParams = jsonObject['bundleParams']

    basedir = os.path.abspath(os.path.dirname(__file__))
    print "Load Hyperflex config from directory: ", basedir
    bf = BundleFactory.loadBundleByID(bundleId, basedir + "/data/")
    bf.extractBundle()
    wb1 = Workbook()
    ws1 = wb1.create_sheet(title="Spec", index=0)
    mapping1 = {BundleFactory.CAT_PN: 2, BundleFactory.CAT_NAME: 3, BundleFactory.CAT_QUANTITY : 4, BundleFactory.CAT_PRICE: 5,
            BundleFactory.CAT_TOTAL: 6}
    headers1 = {BundleFactory.CAT_PN: "Part Number",
            BundleFactory.CAT_QUANTITY: "Qty",
            BundleFactory.CAT_PRICE: "Unit List Price",
            BundleFactory.CAT_TOTAL: "Total list price",
            BundleFactory.CAT_NAME: "Description"}
    bf.create_sheet_header(ws1, 1, mapping1, headers1)
    row = bf.spec_to_worksheet(ws1, 2, 'BASE', mapping1, 1, True)
    row = bf.spec_to_worksheet(ws1, row, 'OPTION', mapping1, bundleParams['servers'], True)
    row = bf.spec_to_worksheet(ws1, row, 'ADDON', mapping1, bundleParams['memory'], True)

    bf.create_sheet_header(ws1, row+2, {BundleFactory.CAT_TOTAL : 6}, {BundleFactory.CAT_TOTAL : "Estimate Total"})
    basedir = os.path.abspath(os.path.dirname(__file__))

    #filename = basedir + "/hyperflex_config/hx-" + str(user_session.getUserID()) + "-" + str(
    #    datetime.datetime.now().microsecond) + ".xslx"

    file_ending = str(user_session.getUserID()) + "-" + str(datetime.datetime.now().microsecond)

    dest_filename = os.path.join(app.config['UPLOAD_FOLDER'], "hx-" + file_ending + ".xlsx")

    print "Filename: ", dest_filename
    wb1.save(filename=dest_filename)

    # Load config to database
    name = "Hyperflex" + file_ending
    # Load specification to database
    try:
        print "Spec name: ", name
        spec_factory = SpecFactory(dest_filename, specName=name)
        print "specification factory created"
        hash = spec_factory.uploadSpecToDatabase(db.session, user_session.getUserID())
        print "Spec hash: '%s', return success" % hash
        user_session.registerEvent(db.session, User_action.UPLOAD_SPEC, request,
                                   data={'filename': dest_filename, 'hash': hash})  # register spec upload
        resp = make_response(jsonify({'result': True, 'hash': hash}), 200)
    except Exception as e:
        print "Exceptions happens: ", e
        resp = make_response({'result': False, 'message': 'Upload failed'}, 404)

    user_session.setCookies(resp)
    return resp


    #return send_file(filename, attachment_filename="hx-" + model + ".xlsx")


## -------------- Server start
if __name__ == '__main__':
    #import logging
    #logging.basicConfig(filename='flask-error.log',level=logging.DEBUG)
    print "App static folder: ", app.static_folder
    print "Database: ", app.config['SQLALCHEMY_DATABASE_URI']
    app.run()

