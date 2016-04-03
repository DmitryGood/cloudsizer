__author__ = 'slash'
from specification.BundleFactory import BundleFactory
from openpyxl import Workbook, worksheet

bf1 = BundleFactory('../data/HyperFlex_1_upload.xlsx')
bf2 = BundleFactory('../data/HyperFlex_2_upload.xlsx')
bf3 = BundleFactory('../data/HyperFlex_3_upload.xlsx')

res1 = bf1.extractBundle()
res2 = bf2.extractBundle()
res3 = bf3.extractBundle()


print "------- Bundle extraction result #1 "
print bf1.bundle
print "------- Bundle extraction result #2 "
print bf2.bundle
print "------- Bundle extraction result #3 "
print bf3.bundle

#print "BASE", res['base']
#print "OPTION", res['option']
#print "ADDON", res['addon']
#print "PARAMETERS", res['parameters']

print "---------"
#print bf.bundle
print "Process bundle for web"

'''
def extract_for_web(bundle):
    bundle_option = bundle.bundle['OPTION']
    result = {}
    for i in range(bundle_option['min'], bundle_option['max']+1):
        #print "Server: ----", i
        option_param= {'price': 0, 'core': 0}
        config = bundle_option['config']
        price = bundle.bundle['BASE']['price'] + bundle_option['price'] * i
        new_conf = {}
        for e in config:
            new_conf[e] = config[e] * i
        cores = int(new_conf['cpu'])
        result[cores] = {
            'servers' : i,
            'price' : price,
            'config' : new_conf,
            'memoryLimit' : bundle.bundle['ADDON']['limit'] * i,
            'memorySize' : bundle.bundle['ADDON']['size'],
            'memoryCost' : bundle.bundle['ADDON']['cost']
        }
        #print "------------" + '\n'
    return result

def compare_is_conf1_less(conf_1, conf_2):
    return (conf_1['price'] < conf_2['price'])

def combine_conf(conf_list):
    result = {}
    # iterate configurations list
    for conf in conf_list:
        # iterate configurations
        for key in conf:
            if not result.has_key(key):             # If key hasn't been in result - put it
                result[key] = conf[key]
            else:                                   # If the same key is in result - decide which one will remain
                print "**** Key conflict", key
                if compare_is_conf1_less(conf[key], result[key]):       # if new is less - put it, else remain existing
                    print "----- >>>> Relace key"
                    result[key] = conf[key]
    return result

print "--------------------"
print extract_for_web(bf1).keys()
print "--------------------"
print extract_for_web(bf2).keys()
print "--------------------"
print extract_for_web(bf3).keys()


print "++++ ====== +++++"
print combine_conf([extract_for_web(bf1), extract_for_web(bf2), extract_for_web(bf3)]).keys()
'''

print "--------------------"
print bf1.extract_for_web().keys()
print "--------------------"
print bf2.extract_for_web().keys()
print "--------------------"
print bf3.extract_for_web().keys()


print "++++ ====== +++++"
print BundleFactory.combine_bundles_config([bf1, bf2, bf3]).keys()
#print combine_conf([extract_for_web(bf1), extract_for_web(bf2), extract_for_web(bf3)]).keys()


mapping1 = {BundleFactory.CAT_PN : 2, BundleFactory.CAT_QUANTITY : 3, BundleFactory.CAT_PRICE : 5, BundleFactory.CAT_TOTAL : 6}
headers1 = {BundleFactory.CAT_PN : "Part Number",
           BundleFactory.CAT_QUANTITY : "Qty",
           BundleFactory.CAT_PRICE : "Unit List Price",
           BundleFactory.CAT_TOTAL : "Total list price"}

mapping2 = {BundleFactory.CAT_PN : 1, BundleFactory.CAT_QUANTITY : 2, BundleFactory.CAT_PRICE : 4}
headers2 = {BundleFactory.CAT_PN : "Part Number",
           BundleFactory.CAT_QUANTITY : "Quantity",
           BundleFactory.CAT_PRICE : "List Price",
           }

wb1= Workbook()
ws1 = wb1.create_sheet(title = "Spec", index=0)

#bf1.create_sheet_header(ws1,3,mapping1,headers1)
#bf1.spec_to_worksheet(ws1, 4, 'BASE', mapping1, 3, True)

bf1.create_sheet_header(ws1, 1, mapping2, headers2)
row = bf1.spec_to_worksheet(ws1, 2, 'BASE', mapping2)
row = bf1.spec_to_worksheet(ws1, row, 'OPTION', mapping2, 5)
row = bf1.spec_to_worksheet(ws1, row, 'ADDON', mapping2, 20)


wb1.save(filename="test_workbook_upload.xlsx")

